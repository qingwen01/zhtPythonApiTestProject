#coding=utf8
import os
import sys
import openpyxl

current_path = os.getcwd() #获取当前文件绝对路径
base_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))  #获取当前文件上上级根目录绝对路径
sys.path.append(base_path)  #将该路径添加到系统变量

class HandleExcel:

    #加载excel
    def get_excel(self):
        data = openpyxl.load_workbook(base_path+"/data/zht.xlsx")
        return data

    #获取所有sheet的内容,index是表序号，如1,2,3....
    def get_sheet_data(self,index=None):
        sheet_name = self.get_excel().sheetnames
        if index ==None:
            index = 0
        data = self.get_excel()[sheet_name[index]]
        return data

    #获取某个单元格的内容
    def get_cell_value(self,row,cols):
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data

    #获取行数
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row


    #获取某一行的内容,row是行号，如1,2,3...
    def get_rows_value(self,row=None):
        row_list = []
        if row == None:
            row = 1
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list


    #获取某一列的数据 cols是列号，如A，B，C....
    def get_columns_value(self,cols=None):
        cols_list = []
        if cols ==None:
            cols = "A"
        for i in self.get_sheet_data()[cols]:
            cols_list.append(i.value)
        return cols_list

    #获取行号，case_id是用例编号，如zht_001,zht_002....,；cols是列号，如A，B，C....
    def get_rows_numble(self,case_id,cols=None):
        num = 1
        for col_data in self.get_columns_value(cols):
            if case_id == col_data:
                return num
            num = num+1
        return num


    #获取表单里面所有数据
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_columns_value(i+2))
        return data_list



    #向Excel某个单元格写入数据
    def excel_write_data(self,row,cols,value):
        wb = self.get_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(base_path+"/data/zht.xlsx")



handleExcel = HandleExcel()

if __name__ == '__main__':
    handleExcel = HandleExcel()
    # print(handleExcel.get_rows_value(7))
    # print(handleExcel.get_sheet_data(2))
    # print(handleExcel.get_columns_value("A"))
    # print(handleExcel.get_rows_numble("zht_005"))
    print(handleExcel.get_excel_data())
